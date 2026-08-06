from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404
from django.shortcuts import redirect, render
from django.shortcuts import get_object_or_404
from urllib.parse import urlencode
from django.urls import reverse
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy
from django.utils import translation
from django.views import generic
from rest_framework import response, generics as rest_generics
from datetime import datetime
from django.core.paginator import Paginator
from django.db.models import Q, QuerySet
import re as re_module
from functools import reduce
from django.db.models import Sum
from django.forms.models import model_to_dict
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
import itertools
from openpyxl.utils import get_column_letter
from datetime import datetime

from process_manager.models import Phase, Activity, Task, Project
from authentication.models import Facilitator
from dashboard.facilitators.forms import FacilitatorForm, FilterTaskForm, UpdateFacilitatorForm, FilterFacilitatorForm
from dashboard.mixins import AJAXRequestMixin, PageMixin, JSONResponseMixin
from no_sql_client import NoSQLClient
import grm_client
from dashboard.utils import (
    sync_geographicalunits_with_cvd_on_facilittor, sync_tasks
)
from authentication.permissions import (
    CDDSpecialistPermissionRequiredMixin, SuperAdminPermissionRequiredMixin,
    AdminPermissionRequiredMixin
    )
from dashboard.facilitators.functions import (
    get_cvds, get_cvd_name_by_village_id, is_village_principal, single_task_by_cvd,
    clear_facilitator_docs_by_administrativelevels_and_save_to_backup_db, 
    get_headquarters_village_id
)
from administrativelevels import models as administrativelevels_models
from assignments.models import AssignAdministrativeLevelToFacilitator
from dashboard.administrative_levels.functions import get_administrative_levels_under_json, get_cascade_villages_by_administrative_level_id
from cdd.functions import datetime_complet_str, exists_id_in_a_dict
from cdd.call_objects_from_other_db import mis_objects_call
from authentication.functions import get_assign_adl_by_facilitatr
from dashboard.tasks import sync_celery_tasks_re
from dashboard.facilitators.views import FacilitatorMixin
from dashboard.administrative_levels.forms import AttachmentFilterForm
from cdd.my_librairies.functions import strip_accents, get_datas_dict
from dashboard.reports.constants import IGNORES, PEULS
from subprojects.models import Project as MisProject
from process_manager.models import AggregatedStatus
from openpyxl import Workbook
from subprojects.models import Subproject



def export_fc_situation_to_excel(request):

    projects_mis = mis_objects_call.filter_objects(MisProject, name__contains='COSO')
    projects_names = [p.name for p in projects_mis]

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"{datetime.now().strftime('%Y%m%d%H%M')}_fc_situations.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 2. Création du classeur et de la feuille de calcul
    wb = Workbook()
    ws = wb.active
    ws.title = "Situations CDD"

    # --- 3. Définition des En-têtes de Colonnes (Headers) ---
    
    columns_lazy = [
        _("FC"), _("CVD initialement affectés"), _("CVD de Stabilisation"), _("Total CVD")
    ] + [f"{c} {p}" for p in projects_names for c in [_("Nombre de tâches achevées"), _("Pourcentage d'achevement"), _("Nombre de tâches validées"), _("Pourcentage de validation"), _("Nombre de tâches")]] + [
        _("Nombre de tâches achevées"), _("Pourcentage d'achevement"), _("Nombre de tâches validées"), _("Pourcentage de validation"), _("Nombre de tâches")
    ]
    
    columns = [str(header) for header in columns_lazy]

    ws.append(columns)

    _headquarters_village_ids = {
        p: set(mis_objects_call.filter_objects(administrativelevels_models.CVD,
            headquarters_village__administrative_levels_projects__name__in=[p]
        ).values_list('headquarters_village_id', flat=True)) for p in projects_names
    }
    all_headquarters_village_ids = set().union(*_headquarters_village_ids.values())
    
    # Récupérer les facilitateurs en une seule requête
    facilitators = Facilitator.objects.filter(facilitator_type='community_facilitator', active=True, develop_mode=False, training_mode=False, projects__name__in=projects_names).distinct()
    
    # Ajout des informations aux niveaux administratifs
    _fcs = []
    for facilitator in facilitators:
        ads_for_fc = {
            k : list(set(([int(ad['id']) for ad in facilitator.administrative_levels if ad['project_name'] == k and str(ad['id']).isdigit()] or []) + (facilitator.stabilization_administrative_ids or [])) & v)
            for k, v in _headquarters_village_ids.items()
        }
        aggregs_dict = {
            k: {
                agg['project__name']: {
                    'total_tasks_completed': agg['total_tasks_completed'] or 0,
                    'total_tasks': agg['total_tasks'] or 0,
                    'total_tasks_validated': agg['total_tasks_validated'] or 0,
                    'total_tasks_invalidated': agg['total_tasks_invalidated'] or 0,
                    'total_tasks_invalidated_review': agg['total_tasks_invalidated_review'] or 0,
                    'total_tasks_invalidated_review_completed': agg['total_tasks_invalidated_review_completed'] or 0,
                    'total_tasks_invalidated_review_in_pending': agg['total_tasks_invalidated_review_in_pending'] or 0,
                    'total_tasks_invalidated_unreview': agg['total_tasks_invalidated_unreview'] or 0,
                    'total_tasks_invalidated_unreview_completed': agg['total_tasks_invalidated_unreview_completed'] or 0,
                    'total_tasks_invalidated_unreview_in_pending': agg['total_tasks_invalidated_unreview_in_pending'] or 0,
                    'total_tasks_waiting_validation': agg['total_tasks_waiting_validation'] or 0
                }
                for agg in AggregatedStatus.objects.filter(
                    administrative_level_id__in=v,
                    project__name=k,
                    task__name__icontains="Soutenir la communauté dans la sélection des priorités par sous-composante (1.1",
                    facilitator=None
                ).values('project__name').annotate(
                    total_tasks_completed=Sum('total_tasks_completed'),
                    total_tasks=Sum('total_tasks'),
                    total_tasks_validated=Sum('total_tasks_validated'),
                    total_tasks_invalidated=Sum('total_tasks_invalidated'),
                    total_tasks_invalidated_review=Sum('total_tasks_invalidated_review'),
                    total_tasks_invalidated_review_completed=Sum('total_tasks_invalidated_review_completed'),
                    total_tasks_invalidated_review_in_pending=Sum('total_tasks_invalidated_review_in_pending'),
                    total_tasks_invalidated_unreview=Sum('total_tasks_invalidated_unreview'),
                    total_tasks_invalidated_unreview_completed=Sum('total_tasks_invalidated_unreview_completed'),
                    total_tasks_invalidated_unreview_in_pending=Sum('total_tasks_invalidated_unreview_in_pending'),
                    total_tasks_waiting_validation=Sum('total_tasks_waiting_validation')
                )
            } for k, v in ads_for_fc.items()
        }
        ranking_by_project = {
            "COSO": 0,
            "FA-COSO": 1
        }
        aggregs_dict = dict(sorted([(k, v) for k, v in aggregs_dict.items()], key=lambda x: ranking_by_project.get(x[0].upper(), 99)))
        
        _total_tasks_completed = 0
        _total_tasks_validated = 0
        _total_tasks = 0
        _fcs = [
            facilitator.name,
            len(list(set(facilitator.administrative_levels_ids or []) & all_headquarters_village_ids)),
            len(list(set(facilitator.stabilization_administrative_ids or []) & all_headquarters_village_ids)),
            len(list(set(sum(list(ads_for_fc.values()), []))))
        ]

        for _project_name in projects_names:
            
            v_ag = aggregs_dict.get(_project_name, {}).get(_project_name)
            
            if v_ag:
                _fcs.append(v_ag['total_tasks_completed'])
                _fcs.append(round((v_ag['total_tasks_completed'] / v_ag['total_tasks']) * 100, 2) if v_ag['total_tasks'] > 0 else 0)
                _total_tasks_completed += v_ag['total_tasks_completed']

                _fcs.append(v_ag['total_tasks_validated'])
                _fcs.append(round((v_ag['total_tasks_validated'] / v_ag['total_tasks']) * 100, 2) if v_ag['total_tasks'] > 0 else 0)
                _total_tasks_validated += v_ag['total_tasks_validated']

                _fcs.append(v_ag['total_tasks'])
                _total_tasks += v_ag['total_tasks']
            else:
                _fcs.extend(["", "", "", "", ""])

        # _t_column_remain = (4 + len(projects_names)*5) - len(_fcs)
        # if _t_column_remain:
        #     print(_t_column_remain)
        #     for c in range(_t_column_remain):
        #         _fcs.append("")
                
        _fcs.append(_total_tasks_completed)
        _fcs.append(round((_total_tasks_completed / _total_tasks) * 100, 2) if _total_tasks > 0 else 0)
        
        _fcs.append(_total_tasks_validated)
        _fcs.append(round((_total_tasks_validated / _total_tasks) * 100, 2) if _total_tasks > 0 else 0)
        
        _fcs.append(_total_tasks)

        ws.append(_fcs)
        

    # Ajustement de la largeur des colonnes (Optionnel) ---
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(response)
    print("ok export")
    return response