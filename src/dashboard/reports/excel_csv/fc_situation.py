"""Génération automatisée du classeur « FC_SITUATION » (ex-`8_1_..._FC_SITUATION.xlsx`).

Remplace, en une seule passe lisant **directement les bases CouchDB des facilitateurs**
(+ `backup_db_facilitators_docs`), l'enchaînement manuel :
`/facilitators/export-situations/` + `/administrative-levels/export-situations/`
+ `/statistics/reports/prorities-pav-pac/` + le notebook `analyse.ipynb` + la fusion Excel.

Le périmètre de tâches est paramétrable (une/plusieurs phases, activités ou tâches). Le volet
« au moins 3 priorités renseignées » (feuilles CVD) ne s'active que si la tâche de priorisation
est dans le périmètre ; pour toute autre tâche les feuilles CVD retombent sur
achèvement / validation pures.

Point d'entrée : `build_fc_situation_workbook(params) -> chemin relatif media/`.
"""
import os
from datetime import datetime
from sys import platform

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from authentication.models import Facilitator
from assignments.models import AssignAdministrativeLevelToFacilitator
from administrativelevels.models import CVD, AdministrativeLevel
from process_manager.models import Cycle, Project, Task
from subprojects.models import Project as MisProject
from no_sql_client import NoSQLClient
import grm_client
from cdd.call_objects_from_other_db import mis_objects_call
from dashboard.administrative_levels.functions import get_cascade_villages_by_administrative_level_id
from dashboard.facilitators.repository.db_facilitator_repository import FacilitatorRepository
from dashboard.facilitators.repository.facilitator_criteria import FacilitatorCriteria
from dashboard.statistics.functions_reports import append_elt, get_cvd_index, get_task_by_task_ids
from dashboard.statistics.utils import normaliser_chaine


# --- Constantes -----------------------------------------------------------------

# sql_id des tâches « Soutenir la communauté dans la sélection des priorités ... » (COSO, FA-COSO, PURS)
PRIORITIZATION_TASK_SQL_IDS = {59, 128, 92}
PRIORITIZATION_TASK_NAMES_NORMALIZED = {
    normaliser_chaine("Soutenir la communauté dans la sélection des priorités par composante (1, 2 et 3) à soumettre à la discussion du CCD."),
    normaliser_chaine("Soutenir la communauté dans la sélection des priorités par sous-composante (1.1, 1.2 et 1.3) à soumettre à la discussion du CCD lors de la réunion cantonale d'arbitrage"),
}

# préfixe de feuille par projet CDD (fallback : nom du projet)
SHEET_PREFIX_BY_PROJECT = {"COSO": "COSO_P", "FA-COSO": "FACOSO", "FA COSO": "FACOSO"}

# ordre des projets dans la feuille FC_SITUATION
PROJECT_RANKING = {"COSO": 0, "FA-COSO": 1, "PURS": 2}

FC_METRIC_LABELS = [
    "Nombre de tâches achevées", "Pourcentage d'achevement",
    "Nombre de tâches validées", "Pourcentage de validation", "Nombre de tâches",
]

CVD_GEO_COLUMNS = ["REGION", "PREFECTURE", "COMMUNE", "CANTON", "CVD", "CVD_ID", "VILLAGES"]
CVD_CONTACT_COLUMNS = [
    "AC Name Stabilization", "AC Phone Stabilization", "AC Email Stabilization",
    "AC Name Initial", "AC Phone Initial", "AC Email Initial",
    "Supervisor Name", "Supervisor Phone", "Supervisor Email",
]

ROLLUP_COLUMNS_PRIORITIES = [
    "Canton",
    "Nombre de CVD",
    "Nombre de CVD pour lesquels la tâche est « achevée » et disposant au moins 3 priorités par les AC dans l’App DCC",
    "Nombre de CVD pour lesquels la tâche est « non achevée » ou ne disposant pas au moins 3 priorités",
    "Nombre de CVD pour lesquels la tâche est « validée » et disposant au moins 3 priorités par les SZ dans l’App DCC",
    "Nombre de CVD pour lesquels la tâche est « non validée » ou ne disposant pas au moins 3 priorités",
]
ROLLUP_COLUMNS_GENERIC = [
    "Canton",
    "Nombre de CVD",
    "Nombre de CVD pour lesquels la tâche est « achevée »",
    "Nombre de CVD pour lesquels la tâche est « non achevée »",
    "Nombre de CVD pour lesquels la tâche est « validée »",
    "Nombre de CVD pour lesquels la tâche est « non validée »",
]


# --- Petits utilitaires -------------------------------------------------------

def _to_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _as_id_list(value):
    """Accepte None / "" / "1,2" / [1, "2", None] -> [1, 2]."""
    if value in (None, "", "None", "null", "undefined"):
        return []
    if isinstance(value, str):
        parts = value.split(",")
    elif isinstance(value, (list, tuple, set)):
        parts = list(value)
    else:
        parts = [value]
    out = []
    for p in parts:
        pid = _to_int(p)
        if pid is not None:
            out.append(pid)
    return out


def _percent(numerator, denominator):
    return round((numerator / denominator) * 100, 2) if denominator else 0


def _style_excel_file(filename):
    """En-têtes en gras + alignement (repris du notebook `analyse.ipynb`)."""
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
    wb.save(filename)


# --- Résolution du périmètre ------------------------------------------------

def _resolve_cdd_projects(params):
    """Liste ordonnée des projets CDD (`process_manager.Project`) à inclure dans le classeur.

    Défaut : l'arbre du projet de session (COSO -> [COSO, FA-COSO]).
    Override : `params['cdd_project_names']` (liste de noms).
    """
    names = params.get("cdd_project_names")
    if names:
        projects = list(Project.objects.filter(name__in=names))
    else:
        session_project_id = params.get("session_project_id")
        base = None
        if session_project_id:
            base = Project.objects.filter(id=session_project_id).first()
        if base is None and params.get("session_project_name"):
            base = Project.objects.filter(name=params["session_project_name"]).first()
        if base is None:
            base = Project.objects.filter(name="COSO").first()
        projects = list(base.build_the_tree_structure()) if base else []
    projects.sort(key=lambda p: PROJECT_RANKING.get(p.name.upper().replace("_", "-"), 99))
    return projects


def _resolve_cycle_couch_ids(projects, params):
    """dict {project.couch_id: cycle.couch_id} — le cycle de même `order` que le cycle de session.

    À défaut d'info de session, on prend le dernier cycle (order le plus élevé) de chaque projet.
    """
    session_cycle_couch_id = params.get("session_cycle_couch_id")
    target_order = None
    if session_cycle_couch_id:
        session_cycle = Cycle.objects.filter(couch_id=session_cycle_couch_id).first()
        if session_cycle:
            target_order = session_cycle.order

    mapping = {}
    for project in projects:
        cycle = None
        if target_order is not None:
            cycle = Cycle.objects.filter(project=project, order=target_order).first()
        if cycle is None:
            cycle = Cycle.objects.filter(project=project).order_by("-order").first()
        if cycle:
            mapping[project.couch_id] = cycle.couch_id
    return mapping


def _resolve_task_scope(projects, params):
    """Renvoie (scope_by_project_name, all_sql_ids, three_priorities_rule).

    scope_by_project_name : {nom_projet: set(sql_id)}.
    Cascade : ids_task sinon ids_activity sinon ids_phase sinon toutes les tâches des projets.
    Les identifiants fournis appartiennent en général à un seul projet (les listes de l'UI sont
    filtrées sur le projet de session) : on étend la sélection aux projets frères par
    correspondance de nom normalisé (tâche / activité / phase) pour couvrir COSO **et** FA-COSO
    en une passe.
    """
    project_ids = [p.id for p in projects]
    ids_task = _as_id_list(params.get("ids_task"))
    ids_activity = _as_id_list(params.get("ids_activity"))
    ids_phase = _as_id_list(params.get("ids_phase"))

    all_tasks = list(
        Task.objects.filter(project_id__in=project_ids).select_related("project", "phase", "activity")
    )

    if ids_task:
        picked = [t for t in all_tasks if t.id in ids_task]
        keys = {normaliser_chaine(t.name) for t in picked}
        selected = [t for t in all_tasks if normaliser_chaine(t.name) in keys]
    elif ids_activity:
        picked = [t for t in all_tasks if t.activity_id in ids_activity]
        keys = {normaliser_chaine(t.activity.name) for t in picked}
        selected = [t for t in all_tasks if normaliser_chaine(t.activity.name) in keys]
    elif ids_phase:
        picked = [t for t in all_tasks if t.phase_id in ids_phase]
        keys = {normaliser_chaine(t.phase.name) for t in picked}
        selected = [t for t in all_tasks if normaliser_chaine(t.phase.name) in keys]
    else:
        selected = all_tasks

    scope_by_project_name = {}
    all_sql_ids = set()
    matched_names_norm = set()
    for task in selected:
        scope_by_project_name.setdefault(task.project.name, set()).add(task.id)
        all_sql_ids.add(task.id)
        matched_names_norm.add(normaliser_chaine(task.name))

    rule = params.get("three_priorities_rule")
    if rule is None:
        rule = bool(all_sql_ids & PRIORITIZATION_TASK_SQL_IDS) or bool(
            matched_names_norm & PRIORITIZATION_TASK_NAMES_NORMALIZED
        )
    return scope_by_project_name, all_sql_ids, bool(rule)


def _get_facilitators(params, project_ids, strict=True):
    """Facilitateurs du périmètre.

    `strict=True` (feuille FC_SITUATION) : community_facilitator actifs, comme
    `export_fc_situation_to_excel`.
    `strict=False` (univers CVD des feuilles 3-8) : tous les facilitateurs non-dev/non-test du
    projet, comme `priorities_pav_pac_situation` — nécessaire pour retrouver les CVD servis par
    des comptes inactifs / de stabilisation.
    """
    project_ids = [str(pid) for pid in project_ids] or None
    facilitator_dbs_name = params.get("facilitator_dbs_name") or []
    if facilitator_dbs_name:
        return Facilitator.objects.filter(
            develop_mode=False, training_mode=False, no_sql_db_name__in=facilitator_dbs_name
        )

    ids_administrativelevel = params.get("ids_administrativelevel")
    if ids_administrativelevel:
        liste_villages = get_cascade_villages_by_administrative_level_id(ids_administrativelevel)
        project_mis = mis_objects_call.filter_objects(
            MisProject, name=params.get("session_project_name")
        ).first()
        project_mis_id = project_mis.id if project_mis else 1
        assign_facilitators = AssignAdministrativeLevelToFacilitator.objects.using("mis").filter(
            administrative_level_id__in=[int(v["administrative_id"]) for v in liste_villages],
            project_id=project_mis_id,
            activated=True,
        )
        criteria = FacilitatorCriteria(
            id__in=list({int(f.facilitator_id) for f in assign_facilitators}),
            develop_mode=False,
            training_mode=False,
            projects__id=project_ids,
            facilitator_type="community_facilitator" if strict else None,
            active=True if strict else None,
        )
    else:
        criteria = FacilitatorCriteria(
            develop_mode=False,
            training_mode=False,
            projects__id=project_ids,
            facilitator_type="community_facilitator" if strict else None,
            active=True if strict else None,
        )
    return FacilitatorRepository().find_by_criteria(criteria=criteria).distinct()


def _headquarters_village_ids_by_project(project_names):
    """{nom_projet: set(headquarters_village_id)} — repris de `export_fc_situation_to_excel`."""
    return {
        name: set(
            mis_objects_call.filter_objects(
                CVD,
                headquarters_village__administrative_levels_projects__name__in=[name],
            ).values_list("headquarters_village_id", flat=True)
        )
        for name in project_names
    }


# --- Lecture CouchDB ---------------------------------------------------------

def _iter_scope_task_docs(db, project_couch_ids, cycle_couch_id_by_project, all_sql_ids):
    """Docs `type=task` du périmètre pour une base CouchDB donnée (une requête par projet)."""
    sql_ids = list(all_sql_ids)
    for project_couch_id in project_couch_ids:
        selector = {"type": "task", "project_id": project_couch_id}
        cycle_couch_id = cycle_couch_id_by_project.get(project_couch_id)
        if cycle_couch_id:
            selector["cycle_id"] = cycle_couch_id
        if sql_ids:
            selector["sql_id"] = {"$in": sql_ids}
        try:
            for row in db.get_query_result(selector):
                yield row
        except Exception as exc:  # noqa: BLE001 - on log et on continue
            print(f"[fc_situation] get_query_result KO ({project_couch_id}): {exc}")


# --- Feuille FC_SITUATION --------------------------------------------------

def _dedup_task_status_by_project(task_docs_by_project):
    """{nom_projet: {(sql_id, adl_id): {'completed', 'validated'}}} — fusion OU sur toutes les copies.

    Les docs proviennent du pool global (toutes les bases FC du périmètre + backup) : la même
    tâche/village peut apparaître dans plusieurs bases (stabilisation), on garde l'état le plus
    avancé.
    """
    status = {}
    for project_name, docs in task_docs_by_project.items():
        per_key = status.setdefault(project_name, {})
        for doc in docs:
            sid = _to_int(doc.get("sql_id"))
            aid = _to_int(doc.get("administrative_level_id"))
            if sid is None or aid is None:
                continue
            entry = per_key.setdefault((sid, aid), {"completed": False, "validated": False})
            if doc.get("completed"):
                entry["completed"] = True
            if doc.get("validated") is True:
                entry["validated"] = True
    return status


def _build_fc_situation_sheet(facilitators, project_names, scope_by_project_name, task_status_by_project):
    hq_ids_by_project = _headquarters_village_ids_by_project(project_names)
    all_hq_ids = set().union(*hq_ids_by_project.values()) if hq_ids_by_project else set()

    columns = ["FC", "CVD initialement affectés", "CVD de Stabilisation", "Total CVD"]
    for name in project_names:
        columns += [f"{label} {name}" for label in FC_METRIC_LABELS]
    columns += list(FC_METRIC_LABELS)

    rows = []
    for facilitator in facilitators:
        ads_for_fc = {
            name: list(
                (
                    {
                        int(ad["id"])
                        for ad in (facilitator.administrative_levels or [])
                        if ad.get("project_name") == name and str(ad.get("id")).isdigit()
                    }
                    | set(facilitator.stabilization_administrative_ids or [])
                )
                & hq_ids
            )
            for name, hq_ids in hq_ids_by_project.items()
        }

        row = [
            facilitator.name,
            len(set(facilitator.administrative_levels_ids or []) & all_hq_ids),
            len(set(facilitator.stabilization_administrative_ids or []) & all_hq_ids),
            len({v for values in ads_for_fc.values() for v in values}),
        ]

        g_completed = g_validated = g_total = 0
        for name in project_names:
            scope_ids = scope_by_project_name.get(name, set())
            fc_ads = set(ads_for_fc.get(name, []))
            entries = [
                v for (sid, aid), v in task_status_by_project.get(name, {}).items()
                if sid in scope_ids and aid in fc_ads
            ]
            nb = len(entries)
            completed = sum(1 for v in entries if v["completed"])
            validated = sum(1 for v in entries if v["validated"])
            if nb == 0:
                # même rendu que `export_fc_situation_to_excel` : bloc projet vide
                row += ["", "", "", "", ""]
            else:
                row += [completed, _percent(completed, nb), validated, _percent(validated, nb), nb]
            g_completed += completed
            g_validated += validated
            g_total += nb

        row += [
            g_completed, _percent(g_completed, g_total),
            g_validated, _percent(g_validated, g_total), g_total,
        ]
        rows.append(row)

    return pd.DataFrame(rows, columns=columns)


# --- Feuilles CVD ---------------------------------------------------------

def _priorities_count(doc):
    """Nombre de priorités du village renseignées dans `form_response` (0 si absent)."""
    form_response = doc.get("form_response") or []
    try:
        sous_composante_11 = dict(list(form_response)[0]).get("sousComposante11") or {}
        return len(sous_composante_11.get("prioritesDuVillage") or [])
    except Exception:  # noqa: BLE001
        return 0


def _collect_cvd_status(task_docs_by_project, three_priorities_rule):
    """{nom_projet: {hq_village_id: {'completed', 'validated', 'nb_prio'}}} agrégé sur les docs."""
    status = {}
    for project_name, docs in task_docs_by_project.items():
        per_cvd = status.setdefault(project_name, {})
        for doc in docs:
            hq_id = _to_int(doc.get("administrative_level_id"))
            if hq_id is None:
                continue
            entry = per_cvd.setdefault(
                hq_id,
                {"completed": True, "validated": True, "nb_prio": 0, "name": doc.get("administrative_level_name")},
            )
            entry["completed"] = entry["completed"] and bool(doc.get("completed"))
            entry["validated"] = entry["validated"] and (doc.get("validated") is True)
            if three_priorities_rule:
                entry["nb_prio"] = max(entry["nb_prio"], _priorities_count(doc))
    return status


def _geo_for_cvd(hq_village_id, cache):
    """(REGION, PREFECTURE, COMMUNE, CANTON, CVD_name, CVD_ID, VILLAGES) pour un village siège.

    Renvoie None si le village siège est introuvable dans `mis` (CVD alors ignoré, comme la
    jointure SQL du notebook).
    """
    if hq_village_id in cache:
        return cache[hq_village_id]
    adl = mis_objects_call.filter_objects(AdministrativeLevel, id=hq_village_id).first()
    result = None
    if adl:
        cvd = adl.cvd
        canton = adl.parent
        commune = canton.parent if canton else None
        prefecture = commune.parent if commune else None
        region = prefecture.parent if prefecture else None
        result = (
            region.name if region else "",
            prefecture.name if prefecture else "",
            commune.name if commune else "",
            canton.name if canton else "",
            (cvd.name if cvd and cvd.name else adl.name),
            cvd.id if cvd else 0,
            ", ".join(o.name for o in cvd.get_villages()) if cvd else adl.name,
        )
    cache[hq_village_id] = result
    return result


def _contacts_map_by_hq_village(hq_village_ids, project_mis_id):
    """{headquarters_village_id: {colonnes contacts}} — porté de `export_administrativelels_situation_to_excel`."""
    contacts = {hq_id: {col: "" for col in CVD_CONTACT_COLUMNS} for hq_id in hq_village_ids}
    if not hq_village_ids:
        return contacts

    # --- AC initial (affectation MIS la plus récente) ---
    assigned = (
        AssignAdministrativeLevelToFacilitator.objects.using("mis")
        .filter(administrative_level_id__in=list(hq_village_ids), project_id=project_mis_id)
        .order_by("-updated_date")
        .values_list("administrative_level_id", "facilitator_id")
    )
    latest_by_adl = {}
    for adl_id, facilitator_id in assigned:
        fid = _to_int(facilitator_id)
        if fid is not None:
            latest_by_adl.setdefault(_to_int(adl_id), fid)
    facilitators = Facilitator.objects.in_bulk(set(latest_by_adl.values()), field_name="id")
    for adl_id, facilitator_id in latest_by_adl.items():
        fac = facilitators.get(facilitator_id)
        if fac and adl_id in contacts:
            contacts[adl_id]["AC Name Initial"] = fac.name or ""
            contacts[adl_id]["AC Phone Initial"] = fac.phone or ""
            contacts[adl_id]["AC Email Initial"] = fac.email or ""

    # --- AC de stabilisation + superviseur (GRM) ---
    try:
        stabilized = grm_client.get_facilitator_by_village(list(hq_village_ids)) or []
    except Exception as exc:  # noqa: BLE001
        print(f"[fc_situation] GRM get_facilitator_by_village KO: {exc}")
        stabilized = []

    for elt in stabilized:
        representative = (elt or {}).get("representative") or {}
        groups = representative.get("groups") or []
        if not representative.get("is_active"):
            continue
        grm_client.attach_administrative_regions_objects(elt)
        covered = set(elt.get("administrative_regions") or [])
        for region_obj in elt.get("administrative_regions_objects") or []:
            covered.update(str(v["id"]) for v in region_obj.get("villages", []))
        for hq_id in hq_village_ids:
            if str(hq_id) not in covered:
                continue
            if "CommunityFacilitator" in groups:
                contacts[hq_id]["AC Name Stabilization"] = representative.get("name", "")
                contacts[hq_id]["AC Phone Stabilization"] = representative.get("phone", "")
                contacts[hq_id]["AC Email Stabilization"] = representative.get("email", "")
            elif "Supervisor" in groups:
                contacts[hq_id]["Supervisor Name"] = representative.get("name", "")
                contacts[hq_id]["Supervisor Phone"] = representative.get("phone", "")
                contacts[hq_id]["Supervisor Email"] = representative.get("email", "")
    return contacts


def _build_cvd_sheets(project, cvd_status, three_priorities_rule, geo_cache):
    """Renvoie {nom_feuille: DataFrame} pour un projet CDD : rollup canton + 2 listes CVD."""
    prefix = SHEET_PREFIX_BY_PROJECT.get(project.name.upper(), project.name)
    per_cvd = cvd_status.get(project.name, {})
    hq_ids = list(per_cvd.keys())

    project_mis = mis_objects_call.filter_objects(MisProject, name=project.name).first()
    contacts = _contacts_map_by_hq_village(hq_ids, project_mis.id if project_mis else 1)

    list_not_3 = []      # tâche non achevée OU < 3 priorités
    list_invalid = []    # < 3 priorités OU non validée
    rollup = {}          # canton -> compteurs

    for hq_id, st in per_cvd.items():
        geo = _geo_for_cvd(hq_id, geo_cache)
        if geo is None:
            continue
        region, prefecture, commune, canton, cvd_name, cvd_id, villages = geo
        has_3 = (st["nb_prio"] >= 3) if three_priorities_rule else True
        completed, validated = st["completed"], st["validated"]

        base_row = {
            "REGION": region, "PREFECTURE": prefecture, "COMMUNE": commune, "CANTON": canton,
            "CVD": cvd_name, "CVD_ID": cvd_id, "VILLAGES": villages,
        }
        base_row.update(contacts.get(hq_id, {col: "" for col in CVD_CONTACT_COLUMNS}))

        if (not completed) or (not has_3):
            list_not_3.append(base_row)
        if (not has_3) or (not validated):
            list_invalid.append(base_row)

        bucket = rollup.setdefault(
            canton or "(sans canton)",
            {"cvd": 0, "done": 0, "not_done": 0, "valid": 0, "not_valid": 0},
        )
        bucket["cvd"] += 1
        if completed and has_3:
            bucket["done"] += 1
        else:
            bucket["not_done"] += 1
        if validated and has_3:
            bucket["valid"] += 1
        else:
            bucket["not_valid"] += 1

    rollup_columns = ROLLUP_COLUMNS_PRIORITIES if three_priorities_rule else ROLLUP_COLUMNS_GENERIC
    rollup_rows = [
        [canton, c["cvd"], c["done"], c["not_done"], c["valid"], c["not_valid"]]
        for canton, c in sorted(rollup.items())
    ]
    totals = [sum(col) for col in zip(*[r[1:] for r in rollup_rows])] if rollup_rows else [0] * 5
    rollup_rows.append(["TOTAL", *totals])

    list_columns = CVD_GEO_COLUMNS + CVD_CONTACT_COLUMNS
    sort_keys = ["REGION", "PREFECTURE", "COMMUNE", "CANTON", "CVD"]

    def _df(records):
        df = pd.DataFrame(records, columns=list_columns)
        if not df.empty:
            df = df.sort_values(by=sort_keys).reset_index(drop=True)
        return df

    if three_priorities_rule:
        name_not_3 = f"{prefix} PAS DES 3 PRIORITES"
        name_invalid = f"{prefix} INVALIDEES OU PAS 3 PRIO"
    else:
        name_not_3 = f"{prefix} NON ACHEVEES"
        name_invalid = f"{prefix} NON VALIDEES"

    return {
        name_invalid: _df(list_invalid),
        name_not_3: _df(list_not_3),
        prefix: pd.DataFrame(rollup_rows, columns=rollup_columns),
    }


# --- Orchestration --------------------------------------------------------

def build_fc_situation_workbook(params):
    """Construit le classeur et renvoie son chemin relatif (sous `media/`).

    Clés attendues dans `params` :
        session_project_id, session_project_name, session_project_couch_id,
        session_cycle_couch_id, cycle_id,
        type, ids_administrativelevel, facilitator_dbs_name,
        ids_phase, ids_activity, ids_task,
        cdd_project_names (optionnel), three_priorities_rule (optionnel, auto sinon).
    """
    nsc = NoSQLClient()

    projects = _resolve_cdd_projects(params)
    if not projects:
        raise ValueError("Aucun projet CDD résolu pour ce périmètre.")
    project_names = [p.name for p in projects]
    project_couch_id_by_name = {p.name: p.couch_id for p in projects}
    cycle_couch_id_by_project = _resolve_cycle_couch_ids(projects, params)

    scope_by_project_name, all_sql_ids, three_priorities_rule = _resolve_task_scope(projects, params)

    project_ids = [p.id for p in projects]
    fc_situation_facilitators = list(_get_facilitators(params, project_ids, strict=True))
    cvd_universe_facilitators = list(_get_facilitators(params, project_ids, strict=False))

    # 1) Pool global des docs de tâches du périmètre : toutes les bases FC (jeu large) + backup.
    #    Sert aussi bien à la feuille FC_SITUATION (comptage par FC via les villages affectés /
    #    de stabilisation) qu'aux feuilles CVD.
    task_docs_by_project = {name: [] for name in project_names}
    project_couch_ids = [project_couch_id_by_name[n] for n in project_names]

    for facilitator in cvd_universe_facilitators:
        try:
            db = nsc.get_db(facilitator.no_sql_db_name)
        except Exception as exc:  # noqa: BLE001
            print(f"[fc_situation] db KO {facilitator.name}: {exc}")
            continue
        for doc in _iter_scope_task_docs(db, project_couch_ids, cycle_couch_id_by_project, all_sql_ids):
            if doc.get("type") == "task" and doc.get("project_name") in task_docs_by_project:
                task_docs_by_project[doc["project_name"]].append(doc)

    try:
        backup_db = nsc.get_db("backup_db_facilitators_docs")
        backup_rows = backup_db.all_docs(include_docs=True)["rows"]
        cycle_couch_ids = set(cycle_couch_id_by_project.values())
        backup_docs = [
            r["doc"] for r in backup_rows
            if r.get("doc")
            and r["doc"].get("project_id") in project_couch_ids
            and (not cycle_couch_ids or r["doc"].get("cycle_id") in cycle_couch_ids)
        ]
        for doc in get_task_by_task_ids(
            [{"doc": d} for d in backup_docs],
            list(all_sql_ids),
            list(PRIORITIZATION_TASK_NAMES_NORMALIZED) if three_priorities_rule else [],
        ):
            if doc.get("project_name") in task_docs_by_project:
                task_docs_by_project[doc["project_name"]].append(doc)
    except Exception as exc:  # noqa: BLE001
        print(f"[fc_situation] backup_db KO: {exc}")

    # 2) Feuille FC_SITUATION (une ligne par FC), comptée sur le pool global
    fc_situation_df = _build_fc_situation_sheet(
        fc_situation_facilitators, project_names, scope_by_project_name,
        _dedup_task_status_by_project(task_docs_by_project),
    )

    cvd_status = _collect_cvd_status(task_docs_by_project, three_priorities_rule)

    # 3) Feuilles CVD par projet
    geo_cache = {}
    invalid_sheets, not3_sheets, rollup_sheets = {}, {}, {}
    for project in projects:
        sheets = _build_cvd_sheets(project, cvd_status, three_priorities_rule, geo_cache)
        for name, df in sheets.items():
            if "INVALID" in name or "NON VALID" in name:
                invalid_sheets[name] = df
            elif "PRIORITES" in name or "NON ACHEV" in name:
                not3_sheets[name] = df
            else:
                rollup_sheets[name] = df

    # 4) Assemblage : ordre calqué sur `8_1_..._FC_SITUATION.xlsx`
    if not os.path.exists("media/utils/exports"):
        os.makedirs("media/utils/exports")
    file_path = (
        f"utils/exports/fc_situation_"
        f"{datetime.today().replace(microsecond=0).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    with pd.ExcelWriter("media/" + file_path) as writer:
        fc_situation_df.to_excel(writer, sheet_name="FC_SITUATION", index=False)
        for name, df in invalid_sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in not3_sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in rollup_sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    try:
        _style_excel_file("media/" + file_path)
    except Exception as exc:  # noqa: BLE001
        print(f"[fc_situation] style KO: {exc}")

    return file_path.replace("/", "\\\\") if platform == "win32" else file_path
