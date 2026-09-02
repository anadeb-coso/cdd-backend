from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404
from django.shortcuts import redirect, render
from django.shortcuts import get_object_or_404
from urllib.parse import urlencode
from django.urls import reverse
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy
from django.utils import translation
from django.views import generic
from rest_framework import response, generics as rest_generics
from datetime import datetime
from django.core.paginator import Paginator
from django.db.models import Q, QuerySet
import re as re_module
from functools import reduce
from django.db.models import Sum
from django.forms.models import model_to_dict
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
import itertools
from openpyxl.utils import get_column_letter
from datetime import datetime

from process_manager.models import Phase, Activity, Task, Project
from authentication.models import Facilitator
from dashboard.facilitators.forms import FacilitatorForm, FilterTaskForm, UpdateFacilitatorForm, FilterFacilitatorForm
from dashboard.mixins import AJAXRequestMixin, PageMixin, JSONResponseMixin
from no_sql_client import NoSQLClient
import grm_client
from dashboard.utils import (
    sync_geographicalunits_with_cvd_on_facilittor, sync_tasks
)
from authentication.permissions import (
    CDDSpecialistPermissionRequiredMixin, SuperAdminPermissionRequiredMixin,
    AdminPermissionRequiredMixin
    )
from dashboard.facilitators.functions import (
    get_cvds, get_cvd_name_by_village_id, is_village_principal, single_task_by_cvd,
    clear_facilitator_docs_by_administrativelevels_and_save_to_backup_db, 
    get_headquarters_village_id
)
from administrativelevels import models as administrativelevels_models
from assignments.models import AssignAdministrativeLevelToFacilitator
from dashboard.administrative_levels.functions import get_administrative_levels_under_json, get_cascade_villages_by_administrative_level_id
from cdd.functions import datetime_complet_str, exists_id_in_a_dict
from cdd.call_objects_from_other_db import mis_objects_call
from authentication.functions import get_assign_adl_by_facilitatr
from dashboard.tasks import sync_celery_tasks_re
from dashboard.facilitators.views import FacilitatorMixin
from dashboard.administrative_levels.forms import AttachmentFilterForm
from cdd.my_librairies.functions import strip_accents, get_datas_dict
from dashboard.reports.constants import IGNORES, PEULS
from subprojects.models import Project as MisProject
from process_manager.models import AggregatedStatus
from openpyxl import Workbook
from subprojects.models import Subproject



def export_administrativelels_situation_to_excel(request):

    id_region = request.GET.get('id_region')
    id_prefecture = request.GET.get('id_prefecture')
    id_commune = request.GET.get('id_commune')
    id_canton = request.GET.get('id_canton')
    id_village = request.GET.get('id_village')
    type_field = request.GET.get('type_field')
    _id = 0

    facilitators_stabilized = []

    # Projet ciblé : override optionnel via ?project=<nom> (sinon projet de session),
    # pour permettre l'export "par projet" (COSO, FA-COSO...) sans changer le projet de session.
    override_project_name = request.GET.get('project') or request.GET.get('project_name')
    if override_project_name:
        _cdd_project = Project.objects.filter(name=override_project_name).first()
        _session_project_name = _cdd_project.name if _cdd_project else override_project_name
        _mis_project = mis_objects_call.filter_objects(MisProject, name=_session_project_name).first()
        _session_project_mis_id = _mis_project.id if _mis_project else request.session.get("project_mis_id")
        _session_project_id = _cdd_project.id if _cdd_project else request.session.get("project_id")
        _first_cycle = _cdd_project.cycle_set.order_by('order').first() if _cdd_project else None
        _session_cycle_id = _first_cycle.id if _first_cycle else request.session.get("cycle_id")
    else:
        _session_project_name = request.session.get('project_name')
        _session_project_mis_id = request.session.get("project_mis_id")
        _session_project_id = request.session.get("project_id")
        _session_cycle_id = request.session.get("cycle_id")

    project_mis = mis_objects_call.filter_objects(MisProject, name=_session_project_name).first()
    project_mis_id = project_mis.id if project_mis else 1


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"{datetime.now().strftime('%Y%m%d%H%M')}_{'coso_parent' if project_mis.name == 'COSO' else str(project_mis.name).lower().replace(' ', '_')}_cycle_cdd_situations.xlsx" #202605201718_coso_parent_cycle_cdd_situations
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 2. Création du classeur et de la feuille de calcul
    wb = Workbook()
    ws = wb.active
    ws.title = "Situations CDD"

    # --- 3. Définition des En-têtes de Colonnes (Headers) ---
    
    columns_lazy = [
        _("Region"), _("Prefecture"), _("Commune"), _("Canton"), _("CVD"), _("CVD_ID"), _("Village"),

        _("A SOUS-PROJET"),

        _("Tasks Total"), _("Tasks Completed"), f'{_("Tasks")} {_("Pending")}', 
        
        _("Tasks Validated"), _("Tasks Invalidated"), _("Tasks Invalidated Review"), _("Tasks Invalidated Review No completed"), 
        _("Tasks Invalidated Unreview"), _("Tasks Waiting Validation"),

        _("Percentage Completed"), _("Percentage of completed tasks validated"),

        _("Last Activity CDD"),

        _("AC Name Stabilization"), _("AC Phone Stabilization"), _("AC Email Stabilization"), 
        _("AC Name Initial"), _("AC Phone Initial"), _("AC Email Initial"), 

        _("Supervisor Name"), _("Supervisor Phone"), _("Supervisor Email"),
    ]
    
    columns = [str(header) for header in columns_lazy]

    ws.append(columns)

    if (id_region or id_prefecture or id_commune or id_canton or id_village) and type_field:
        _type = None
        if id_region and type_field == "region":
            _type = "region"
            _id = id_region
        elif id_prefecture and type_field == "prefecture":
            _type = "prefecture"
            _id = id_prefecture
        elif id_commune and type_field == "commune":
            _type = "commune"
            _id = id_commune
        elif id_canton and type_field == "canton":
            _type = "canton"
            _id = id_canton
        elif id_village and type_field == "village":
            _type = "village"
            _id = id_village

        nsc = NoSQLClient()
        liste_villages = []
        
        liste_villages = get_cascade_villages_by_administrative_level_id(_id)

        if type(_id) is not list:
            # assign_facilitators = assign_facilitators.filter(
            #     administrative_level_id__in=[int(v['administrative_id']) for v in liste_villages]
            # )
            administrativelels = mis_objects_call.filter_objects(administrativelevels_models.CVD,
                # headquarters_village__id__in=list(set([int(f.administrative_level_id) for f in assign_facilitators]))
                headquarters_village__id__in=[adl.id for adl in project_mis.administrative_levels.filter(id__in=[int(v['administrative_id']) for v in liste_villages])]
            )
        else:
            administrativelels = mis_objects_call.filter_objects(administrativelevels_models.CVD,
                # headquarters_village__id__in=list(set([int(f.administrative_level_id) for f in assign_facilitators]))
                headquarters_village__id__in=[adl.id for adl in project_mis.administrative_levels.all()]
            )
        
    else:
        is_training = bool(request.GET.get('is_training', "False") == "True")
        is_develop = bool(request.GET.get('is_develop', "False") == "True")
        administrativelels = mis_objects_call.filter_objects(administrativelevels_models.CVD,
            # headquarters_village__id__in=list(set([int(f.administrative_level_id) for f in assign_facilitators]))
            headquarters_village__id__in=[adl.id for adl in project_mis.administrative_levels.all()]
        )

    # Récupérer les IDs des villages du quartier général
    _headquarters_village_ids = [adl_cvd.headquarters_village_id for adl_cvd in administrativelels]

    # Sous-requête pour obtenir la dernière affectation désactivée
    last_deactivated_subquery = mis_objects_call.filter_objects(
        AssignAdministrativeLevelToFacilitator,
        administrative_level_id=OuterRef('administrative_level_id'),
        project_id=_session_project_mis_id,
        activated=False
    ).order_by('-updated_date').values('facilitator_id')[:1]  # Prendre le plus récent

    # Requête principale : récupérer les affectations actives ou la dernière désactivée si aucune active n'existe
    _assigned_data = mis_objects_call.filter_objects(
        AssignAdministrativeLevelToFacilitator,
        administrative_level_id__in=_headquarters_village_ids,
        project_id=_session_project_mis_id
    ).annotate(
        last_deactivated_facilitator=Subquery(last_deactivated_subquery)
    ).values_list('administrative_level_id', 'facilitator_id', 'last_deactivated_facilitator')

    # Construire la liste finale avec priorité aux affectations actives
    final_assignments = [
        (adl_id, facilitator_id if facilitator_id else last_deactivated_facilitator)
        for adl_id, facilitator_id, last_deactivated_facilitator in _assigned_data
    ]

    # Création des dictionnaires pour accès rapide
    _administrative_level_ids_assigned = {adl_f[0] for adl_f in final_assignments}
    _facilitator_ids_assigned = {int(adl_f[1]) for adl_f in final_assignments}

    # Récupérer les facilitateurs en une seule requête
    facilitators = Facilitator.objects.filter(id__in=_facilitator_ids_assigned).in_bulk(field_name='id')
    
    # Création d'un dictionnaire des affectations
    facilitator_assignments = {
        adl_id: {
            'facilitator_id': int(fac_id),
            'facilitator':  model_to_dict(facilitators.get(int(fac_id))) if facilitators.get(int(fac_id)) else {}
        }
        for adl_id, fac_id in final_assignments
    }

    # Facilitators stabilization data
    facilitators_stabilized = grm_client.get_facilitator_by_village(_headquarters_village_ids)
    
    print("ok facilitators Initial")
    _f_s = {}
    _supervisors_s = {}
    if facilitators_stabilized:
        for elt in [row for row in facilitators_stabilized[:] if (_ for _ in ["CommunityFacilitator", "Supervisor"] if _ in row["representative"]["groups"]) and row["representative"]["is_active"] == True]:
            if elt not in list(_f_s.values()):
                grm_client.attach_administrative_regions_objects(elt)
                for adl_id in _headquarters_village_ids:
                    administratives_stabilized = elt['administrative_regions']
                    administrative_regions_objects = elt.get('administrative_regions_objects')
                    administratives_stabilized = list(set(
                        (administratives_stabilized if administratives_stabilized else []) + list(itertools.chain(*[[str(v['id']) for v in ad['villages']] for ad in (administrative_regions_objects if administrative_regions_objects else [])]))
                    ))
                    if str(adl_id) in administratives_stabilized:
                        if "CommunityFacilitator" in elt["representative"]["groups"]:
                            _f_s[int(adl_id)] = elt
                        elif "Supervisor" in elt["representative"]["groups"]:
                            _supervisors_s[int(adl_id)] = elt
    print("ok facilitators stabilized and supervisors")

    # Récupérer les agrégations associées aux quartiers généraux des villages
    aggregs_by_project = AggregatedStatus.objects.filter(
        administrative_level_id__in=_headquarters_village_ids,
        project_id=_session_project_id,
        cycle_id=_session_cycle_id,
        task=None,
        facilitator=None
    )

    # Agréger les tâches et stocker les résultats sous forme de dictionnaire
    aggregs_dict = {
        agg['administrative_level_id']: {
            'total_tasks_completed': agg['total_tasks_completed'] or 0,
            'total_tasks': agg['total_tasks'] or 0,
            'total_tasks_validated': agg['total_tasks_validated'] or 0,
            'total_tasks_invalidated': agg['total_tasks_invalidated'] or 0,
            'total_tasks_invalidated_review': agg['total_tasks_invalidated_review'] or 0,
            'total_tasks_invalidated_review_completed': agg['total_tasks_invalidated_review_completed'] or 0,
            'total_tasks_invalidated_review_in_pending': agg['total_tasks_invalidated_review_in_pending'] or 0,
            'total_tasks_invalidated_unreview': agg['total_tasks_invalidated_unreview'] or 0,
            'total_tasks_invalidated_unreview_completed': agg['total_tasks_invalidated_unreview_completed'] or 0,
            'total_tasks_invalidated_unreview_in_pending': agg['total_tasks_invalidated_unreview_in_pending'] or 0,
            'total_tasks_waiting_validation': agg['total_tasks_waiting_validation'] or 0
        }
        for agg in aggregs_by_project.values('administrative_level_id').annotate(
            total_tasks_completed=Sum('total_tasks_completed'),
            total_tasks=Sum('total_tasks'),
            total_tasks_validated=Sum('total_tasks_validated'),
            total_tasks_invalidated=Sum('total_tasks_invalidated'),
            total_tasks_invalidated_review=Sum('total_tasks_invalidated_review'),
            total_tasks_invalidated_review_completed=Sum('total_tasks_invalidated_review_completed'),
            total_tasks_invalidated_review_in_pending=Sum('total_tasks_invalidated_review_in_pending'),
            total_tasks_invalidated_unreview=Sum('total_tasks_invalidated_unreview'),
            total_tasks_invalidated_unreview_completed=Sum('total_tasks_invalidated_unreview_completed'),
            total_tasks_invalidated_unreview_in_pending=Sum('total_tasks_invalidated_unreview_in_pending'),
            total_tasks_waiting_validation=Sum('total_tasks_waiting_validation')
        )
    }

    # Récupérer les dernières activités des villages
    latest_activities = {
        agg.administrative_level_id: agg.last_activity
        for agg in aggregs_by_project.order_by('administrative_level_id', '-last_activity')
    }
    print("ok aggregs dict and latest activities")

    # Ajout des informations aux niveaux administratifs
    _administrativelels = []
    for _adl in administrativelels:
        hq_village_id = _adl.headquarters_village_id

        # Vérifier si un facilitateur est assigné
        if hq_village_id in _administrative_level_ids_assigned:
            _adl.is_facilitator_on_this_cvd = True
            facilitator_data = facilitator_assignments.get(hq_village_id)
            _adl.last_facilitator = facilitator_data['facilitator'] if facilitator_data and facilitator_data['facilitator'] else {}
            
        # Vérifier si une agrégation existe pour ce village
        if hq_village_id in aggregs_dict:
            totals = aggregs_dict[hq_village_id]
            _adl.percent_cdd = round((totals['total_tasks_completed'] / totals['total_tasks']) * 100, 2) if totals['total_tasks'] > 0 else 0
            _adl.percent_cdd_validated = round((totals['total_tasks_validated'] / totals['total_tasks_completed']) * 100, 2) if totals['total_tasks_completed'] > 0 else 0

            # Récupérer la dernière activité enregistrée
            _adl.last_activity_cdd = latest_activities.get(hq_village_id)

            _adl.total_tasks_completed = totals['total_tasks_completed']
            _adl.total_tasks_pending = totals['total_tasks'] - totals['total_tasks_completed']
            _adl.total_tasks = totals['total_tasks']
            _adl.total_tasks_validated = totals['total_tasks_validated']
            _adl.total_tasks_invalidated = totals['total_tasks_invalidated']
            _adl.total_tasks_invalidated_review = totals['total_tasks_invalidated_review']
            _adl.total_tasks_invalidated_review_completed = totals['total_tasks_invalidated_review_completed']
            _adl.total_tasks_invalidated_review_in_pending = totals['total_tasks_invalidated_review_in_pending']
            _adl.total_tasks_invalidated_unreview = totals['total_tasks_invalidated_unreview']
            _adl.total_tasks_invalidated_unreview_completed = totals['total_tasks_invalidated_unreview_completed']
            _adl.total_tasks_invalidated_unreview_in_pending = totals['total_tasks_invalidated_unreview_in_pending']
            _adl.total_tasks_waiting_validation = totals['total_tasks_waiting_validation'] + totals['total_tasks_invalidated_review_completed']

        _administrativelel = [
            _adl.get_canton().parent.parent.parent.name if _adl.get_canton() and _adl.get_canton().parent and _adl.get_canton().parent.parent and _adl.get_canton().parent.parent.parent else "",
            _adl.get_canton().parent.parent.name if _adl.get_canton() and _adl.get_canton().parent and _adl.get_canton().parent.parent else "",
            _adl.get_canton().parent.name if _adl.get_canton() and _adl.get_canton().parent else "",
            _adl.get_canton().name if _adl.get_canton() else "",
            _adl.name if _adl.name else "",
            _adl.id,
            _adl.get_names(),

            "OUI" if mis_objects_call.filter_objects(Subproject,
                Q(location_subproject_realized__id__in=[_v.id for _v in _adl.get_villages()]) | Q(cvd__id=_adl.id)
            ).filter(projects__in=[project_mis.id]).exists() else "NON",

            _adl.total_tasks,
            _adl.total_tasks_completed,
            _adl.total_tasks_pending,
            
            _adl.total_tasks_validated,
            _adl.total_tasks_invalidated,
            _adl.total_tasks_invalidated_review,
            _adl.total_tasks_invalidated_review_in_pending,
            _adl.total_tasks_invalidated_unreview,
            _adl.total_tasks_waiting_validation,

            _adl.percent_cdd,
            _adl.percent_cdd_validated,
            
            _adl.last_activity_cdd.replace(tzinfo=None) if _adl.last_activity_cdd else "",

            # Facilitator stabilization data
            (_f_s[_adl.headquarters_village_id]['representative']['name'] if (_f_s and _f_s.get(_adl.headquarters_village_id) and 'representative' in _f_s[_adl.headquarters_village_id] and 'name' in _f_s[_adl.headquarters_village_id]['representative']) else ""),
            (_f_s[_adl.headquarters_village_id]['representative']['phone'] if (_f_s and _f_s.get(_adl.headquarters_village_id) and 'representative' in _f_s[_adl.headquarters_village_id] and 'phone' in _f_s[_adl.headquarters_village_id]['representative']) else ""),
            (_f_s[_adl.headquarters_village_id]['representative']['email'] if (_f_s and _f_s.get(_adl.headquarters_village_id) and 'representative' in _f_s[_adl.headquarters_village_id] and 'email' in _f_s[_adl.headquarters_village_id]['representative']) else ""),

            # Facilitator initial data
            _adl.last_facilitator.get('name', ""),
            _adl.last_facilitator.get('phone', ""),
            _adl.last_facilitator.get('email', ""),

            # Supervisor data
            (_supervisors_s[_adl.headquarters_village_id]['representative']['name'] if (_supervisors_s and _supervisors_s.get(_adl.headquarters_village_id) and 'representative' in _supervisors_s[_adl.headquarters_village_id] and 'name' in _supervisors_s[_adl.headquarters_village_id]['representative']) else ""),
            (_supervisors_s[_adl.headquarters_village_id]['representative']['phone'] if (_supervisors_s and _supervisors_s.get(_adl.headquarters_village_id) and 'representative' in _supervisors_s[_adl.headquarters_village_id] and 'phone' in _supervisors_s[_adl.headquarters_village_id]['representative']) else ""),
            (_supervisors_s[_adl.headquarters_village_id]['representative']['email'] if (_supervisors_s and _supervisors_s.get(_adl.headquarters_village_id) and 'representative' in _supervisors_s[_adl.headquarters_village_id] and 'email' in _supervisors_s[_adl.headquarters_village_id]['representative']) else ""),

        ]
        ws.append(_administrativelel)
    print("ok administrativelels loop")

    # Ajustement de la largeur des colonnes (Optionnel) ---
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(response)
    print("ok export")
    return response